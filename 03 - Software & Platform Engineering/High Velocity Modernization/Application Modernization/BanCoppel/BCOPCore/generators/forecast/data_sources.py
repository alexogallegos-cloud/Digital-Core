"""
data_sources — Loaders extensibles de las fuentes de datos + merge con dedup.

Cada fuente declara su archivo, su loader y una prioridad (mayor gana en fechas solapadas).
Al llegar datos nuevos, agrega una entrada a SOURCES; el pipeline la incorpora sin mas cambios.

Contrato del loader: recibe la ruta absoluta al Excel y devuelve una lista de dicts
{date: pd.Timestamp, eglobal: float, spei: float}.
"""

from dataclasses import dataclass
from typing import Callable
import pandas as pd
import openpyxl


@dataclass
class Source:
    name: str
    relpath: str          # relativo a BCOPCore/
    loader: Callable      # (abs_path) -> list[dict]
    priority: int         # mayor gana en fechas solapadas


# ── loader: 2025 minuto-a-minuto (dos hojas) -> diario ──────────────────────────
def load_minxmin(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def agg(sheet_name):
        ws = wb[sheet_name]
        by_d = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            fecha, _hora, txn = row[0], row[1], row[2]
            if fecha is None:
                continue
            d = fecha.date() if hasattr(fecha, "date") else fecha
            try:
                v = float(txn) if txn is not None else 0.0
            except (TypeError, ValueError):
                v = 0.0
            by_d[d] = by_d.get(d, 0.0) + v
        return by_d

    eg = agg("Eglobal minxmin")
    sp = agg("SPEI Recibidos minxmin")
    wb.close()
    all_d = sorted(set(eg) | set(sp))
    return [{"date": pd.Timestamp(d), "eglobal": eg.get(d, 0.0), "spei": sp.get(d, 0.0)}
            for d in all_d]


# ── loader: 2026 diario preagregado ─────────────────────────────────────────────
def load_diario(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Diario SPEI y Eglobal"]
    recs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            continue
        d = row[0].date() if hasattr(row[0], "date") else row[0]
        recs.append({
            "date": pd.Timestamp(d),
            "eglobal": float(row[6]) if row[6] is not None else 0.0,
            "spei":    float(row[4]) if row[4] is not None else 0.0,
        })
    wb.close()
    return recs


# ── registro de fuentes (agregar nuevas aqui) ────────────────────────────────────
SOURCES = [
    Source("2025-minxmin", "source/spei-aut-ent/Master_Transacciones minxmin_01-01-25 a 4-mar-26.xlsx",
           load_minxmin, priority=1),
    Source("2026-diario",  "source/spei-aut-ent/Transacciones_maestro_Medios_de_Pago.xlsx",
           load_diario, priority=2),   # gana sobre minxmin en ene-mar 2026
]


def load_all(root):
    """Carga todas las fuentes y devuelve un DataFrame diario con dedup por prioridad."""
    frames = []
    for s in sorted(SOURCES, key=lambda x: x.priority):
        recs = s.loader(str(root / s.relpath))
        df = pd.DataFrame(recs)
        df["_prio"] = s.priority
        df["_src"] = s.name
        frames.append(df)
        print(f"  [{s.name}] {len(df)} dias ({df.date.min().date()} a {df.date.max().date()})")
    combined = pd.concat(frames).sort_values(["date", "_prio"])
    combined = combined.drop_duplicates("date", keep="last").sort_values("date").reset_index(drop=True)
    print(f"  merge: {len(combined)} dias ({combined.date.min().date()} a {combined.date.max().date()})")
    return combined
