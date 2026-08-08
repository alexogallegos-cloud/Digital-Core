"""
capacity — Percentiles de carga (txn/min) mes a mes + deteccion de rafagas.

Complementa el forecast de VOLUMEN (diario) con el analisis de CARGA por minuto, para
dimensionar el sistema target y ver si la carga soportada crece.

Inferencia propia (no se adopta la convencion del cliente de percentil sobre 1440 min):
la madrugada muerta contamina el P70. Se infiere la VENTANA OPERATIVA de los datos
(perfil intradia) y los percentiles se calculan sobre la carga de negocio.

Distingue tres cosas:
  - Carga tipica (P70) y pico tipico (P90/P99): demanda servida en horas de operacion.
  - Pico maximo diario: la demanda instantanea maxima de un dia normal.
  - Rafagas: minutos con val >> P99 del dia (colas represadas que se liberan tras un
    valle/incidente). No son carga sostenida; se reportan aparte.
"""

import numpy as np
from collections import defaultdict

# Fuentes minuto-a-minuto por canal: (relpath, kind, param)
#   kind='sheet2025' -> hoja (Fecha, Hora, Transacciones)
#   kind='col2026'   -> hoja 'Min a min SPEI y Eglobal', col con el valor
DIR = "source/spei-aut-ent"
F25 = f"{DIR}/Master_Transacciones minxmin_01-01-25 a 4-mar-26.xlsx"
F26 = f"{DIR}/Transacciones_maestro_Medios_de_Pago.xlsx"

MINUTE_SOURCES = {
    "spei":    [(F25, "sheet2025", "SPEI Recibidos minxmin"), (F26, "col2026", 5)],
    "eglobal": [(F25, "sheet2025", "Eglobal minxmin"),        (F26, "col2026", 7)],
}

DOW = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]


def load_minute_channel(root, channel):
    """Devuelve by_day[date] = list[(minute_of_day, val)] fusionando fuentes (2026 gana)."""
    import openpyxl
    by_day = {}
    for relpath, kind, param in MINUTE_SOURCES[channel]:
        path = str(root / relpath)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        tmp = defaultdict(list)
        if kind == "sheet2025":
            ws = wb[param]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or row[0] is None:
                    continue
                d = row[0].date() if hasattr(row[0], "date") else row[0]
                h = row[1]
                mod = h.hour * 60 + h.minute if hasattr(h, "hour") else 0
                v = row[2]
                tmp[d].append((mod, float(v) if v is not None else 0.0))
        else:  # col2026
            ws = wb["Min a min SPEI y Eglobal"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or row[0] is None:
                    continue
                dt = row[0]
                d = dt.date() if hasattr(dt, "date") else dt
                mod = dt.hour * 60 + dt.minute if hasattr(dt, "hour") else 0
                v = row[param]
                tmp[d].append((mod, float(v) if v is not None else 0.0))
        wb.close()
        for d, vals in tmp.items():
            by_day[d] = vals          # fuente posterior (2026) gana en solape
    return by_day


def intraday_profile(by_day, min_obs=1400):
    acc = np.zeros(1440); cnt = np.zeros(1440)
    for d, mins in by_day.items():
        if len(mins) < min_obs:
            continue
        for mod, v in mins:
            if 0 <= mod < 1440:
                acc[mod] += v; cnt[mod] += 1
    cnt[cnt == 0] = 1
    return acc / cnt


def operating_window(profile, frac=0.25):
    """Minutos cuyo promedio >= frac*pico. Devuelve (set_minutos, (hora0, hora1))."""
    active = np.where(profile >= frac * profile.max())[0]
    return set(int(m) for m in active), (int(active.min()) // 60, int(active.max()) // 60)


def daily_metrics(by_day, win, min_obs=1400):
    """Por dia completo: percentiles en ventana operativa + max + p50 (para deteccion de valle)."""
    out = {}
    for d, mins in by_day.items():
        if len(mins) < min_obs:
            continue
        oper = np.array([v for mod, v in mins if mod in win], float)
        if len(oper) < 30:
            continue
        out[d] = {
            "p70": float(np.percentile(oper, 70)), "p90": float(np.percentile(oper, 90)),
            "p99": float(np.percentile(oper, 99)), "max": float(oper.max()),
            "p50": float(np.percentile(oper, 50)),
        }
    return out


def monthly(daily):
    by_m = defaultdict(list)
    for d, r in daily.items():
        by_m[(d.year, d.month)].append((d, r))
    rows = []
    for (yr, mo) in sorted(by_m):
        rs = [r for _, r in by_m[(yr, mo)]]
        pico_d, pico_v = max(by_m[(yr, mo)], key=lambda t: t[1]["max"])[0], max(r["max"] for r in rs)
        rows.append({
            "m": f"{yr}-{mo:02d}", "year": yr, "month": mo, "days": len(rs),
            "p70": round(float(np.median([r["p70"] for r in rs]))),
            "p90": round(float(np.median([r["p90"] for r in rs]))),
            "p99": round(float(np.median([r["p99"] for r in rs]))),
            "max_med": round(float(np.median([r["max"] for r in rs]))),
            "pico": round(pico_v), "pico_date": str(pico_d),
        })
    return rows


def intraday_profiles(root, cal, bins=288):
    """Perfil intradia NORMALIZADO (suma=1) por canal y tipo de dia (habil/finde), en `bins`
    intervalos (288 = cada 5 min). Reconstruye la curva de cualquier dia: txn/min[bin] =
    perfil[bin] * volumen_diario / (1440/bins)."""
    import numpy as np
    step = 1440 // bins
    out = {}
    for ch in ("eglobal", "spei"):
        by_day = load_minute_channel(root, ch)
        out[ch] = {}
        for tipo, cond in (("habil", lambda d: cal.is_business_day(d)),
                           ("finde", lambda d: not cal.is_business_day(d))):
            acc = np.zeros(bins); n = 0
            for d, mins in by_day.items():
                if len(mins) < 1400 or not cond(d):
                    continue
                b = np.zeros(bins)
                for mod, v in mins:
                    if 0 <= mod < 1440:
                        b[mod // step] += v
                tot = b.sum()
                if tot < 100_000:
                    continue
                acc += b / tot; n += 1
            out[ch][tipo] = (acc / max(n, 1)).tolist()
    return out


def correlated_percentiles(root, cal, d0, d1, w=1, op=(12, 23), top_n=5, _egm=None, _spm=None):
    """
    CALCULO DE PERCENTILES CORRELACIONADOS.
    Carga que SPEI y Autorizador ejercen SIMULTANEAMENTE sobre Informix (recurso compartido),
    en ventanas de `w` minutos: con w>1 mide carga sostenida (promedia y suaviza rafagas de 1 min
    de las que el sistema se restablece); con w=1 es la resolucion cruda por minuto (pico
    instantaneo, sin suavizado ni buffer). Devuelve, para el periodo [d0, d1], TODOS los dias
    (habiles y no habiles —
    SPEI y el Autorizador operan el fin de semana) y horario operativo `op`:
      - P70/P90 por canal (umbral de alerta) y su suma
      - zona de riesgo = % de ventanas con AMBOS >= su P70; incidencia = AMBOS >= su P90
      - correlacion minuto-ventana entre canales
      - top_n de concurrencia sostenida (ambos >= P70) sin caida cerca (capacidad demostrada)
    """
    from collections import defaultdict
    import numpy as np
    if _egm is None:
        _egm = {(d, m): v for d, mm in load_minute_channel(root, "eglobal").items()
                if len(mm) >= 1400 for m, v in mm}
    if _spm is None:
        _spm = {(d, m): v for d, mm in load_minute_channel(root, "spei").items()
                if len(mm) >= 1400 for m, v in mm}
    egm, spm = _egm, _spm

    def ventanas(dic):
        acc = defaultdict(list)
        for (d, m), v in dic.items():
            if d0 <= d <= d1 and op[0]*60 <= m < op[1]*60:   # todos los dias (habiles y no habiles)
                acc[(d, m // w)].append(v)
        return {k: float(np.mean(vs)) for k, vs in acc.items() if len(vs) == w}

    veg, vsp = ventanas(egm), ventanas(spm)
    keys = sorted(veg.keys() & vsp.keys())
    a_eg = np.array([veg[k] for k in keys]); a_sp = np.array([vsp[k] for k in keys])
    com = a_eg + a_sp
    p70e, p90e = np.percentile(a_eg, 70), np.percentile(a_eg, 90)
    p70s, p90s = np.percentile(a_sp, 70), np.percentile(a_sp, 90)
    from scipy.stats import pearsonr
    r = float(pearsonr(a_eg, a_sp)[0]) if len(keys) > 2 else float("nan")
    riesgo = float(np.mean((a_eg >= p70e) & (a_sp >= p70s)))
    incidencia = float(np.mean((a_eg >= p90e) & (a_sp >= p90s)))

    # top_n de concurrencia sostenida (ambos >= P70) sin caida cerca
    comk = {k: com[i] for i, k in enumerate(keys)}
    por_dia = defaultdict(list)
    for k, v in comk.items():
        por_dia[k[0]].append(v)
    med = {d: np.median(vs) for d, vs in por_dia.items()}
    def sin_caida(d, w5):
        return all(comk.get((d, ww), med[d]) >= 0.20 * med[d]
                   for ww in range(w5 - 6, w5 + 7) if (d, ww) in comk)
    conc = [(k, veg[k], vsp[k], comk[k]) for k in keys
            if veg[k] >= p70e and vsp[k] >= p70s and sin_caida(k[0], k[1])]
    conc.sort(key=lambda x: -x[3])
    top, vistos = [], set()
    for (d, w5), e, s, c in conc:
        if d in vistos:
            continue
        vistos.add(d)
        top.append({"fecha": str(d), "hora": f"{(w5*w)//60:02d}:{(w5*w)%60:02d}",
                    "eglobal": round(e), "spei": round(s), "combinada": round(c)})
        if len(top) == top_n:
            break
    prom = {"eglobal": round(np.mean([t["eglobal"] for t in top])) if top else 0,
            "spei": round(np.mean([t["spei"] for t in top])) if top else 0,
            "combinada": round(np.mean([t["combinada"] for t in top])) if top else 0}
    return {
        "periodo": f"{d0} a {d1}", "ventana_min": w, "n_ventanas": len(keys),
        "p70": {"eglobal": round(p70e), "spei": round(p70s), "suma": round(p70e + p70s)},
        "p90": {"eglobal": round(p90e), "spei": round(p90s), "suma": round(p90e + p90s)},
        "correlacion": round(r, 3),
        "pct_zona_riesgo": round(riesgo * 100, 1),
        "pct_incidencia": round(incidencia * 100, 1),
        "top_concurrencia": top, "top_promedio": prom,
    }


def detect_bursts(by_day, daily, win, cal=None, atypical=None, k=2.5, top=20):
    """Minutos con val > k * P99_dia. Marca si hay valle en los 30 min previos (recuperacion)."""
    atypical = atypical or set()
    bursts = []
    for d, mins in by_day.items():
        if d not in daily:
            continue
        p99 = daily[d]["p99"]; p50 = daily[d]["p50"]
        thr = k * p99
        mp = {mod: v for mod, v in mins}
        for mod, v in mins:
            if mod not in win or v <= thr:
                continue
            prev = [mp.get(m, 0) for m in range(max(0, mod - 30), mod)]
            valley = bool(prev) and min(prev) < 0.10 * p50   # colapso previo -> recuperacion
            near_inc = any(abs((d - a).days) <= 1 for a in atypical)
            tag = []
            if cal is not None:
                if d in cal.q15 or d in cal.qfin:
                    tag.append("quincena")
                if not cal.is_business_day(d):
                    tag.append("no-habil")
            bursts.append({
                "date": str(d), "dow": DOW[d.weekday()],
                "time": f"{mod//60:02d}:{mod%60:02d}", "val": round(v),
                "x_over_p99": round(v / p99, 1), "recovery": valley,
                "near_incident": near_inc, "calendar": tag,
            })
    bursts.sort(key=lambda b: -b["val"])
    return bursts[:top]
